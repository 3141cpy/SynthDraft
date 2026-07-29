"""明细栏（BOM）导出 + 装配图导出（Task 10.5）。

输出物：
1. BOM（明细栏）：
   - CSV / JSON / Excel（xlsx 可选，依赖 openpyxl）
   - 字段：件号 / 图号 / 名称 / 数量 / 材料 / 单件质量 / 总质量 / 备注
   - 遵循 GB/T 4458.2《明细栏》格式

2. 装配图（DXF）：
   - 简化装配示意图（基于零件 AABB 绘制矩形 + 标注件号）
   - 标题栏（含装配体名 / 版本 / 日期 / 作者）
   - 明细栏（与 BOM 一致）
   - 遵循 GB/T 4457 / GB/T 4458 / GB/T 18229 基础规则

3. SLDASM：
   - 由 writer.py 通过 SolidWorks API 生成（Windows 平台）
   - 跨平台降级：Linux 输出 STEP 装配体（通过 pythonOCC 或 CadQuery 装配）

设计原则：
- BOM/装配图导出不依赖 SolidWorks（跨平台）
- SLDASM 生成委托给 writer.py，本模块仅负责编排
- 单位：mm
"""

from __future__ import annotations

import csv
import io
import json
import time
from pathlib import Path
from typing import Any

from app.logging import get_logger
from app.schemas.assembly import AssemblySpec, TypedPart

log = get_logger(__name__)


# ===== BOM 导出 =====


def export_bom(
    spec: AssemblySpec,
    output_path: Path,
    format: str = "csv",
) -> Path:
    """导出明细栏（BOM）。

    Args:
        spec: 装配规范
        output_path: 输出文件路径（扩展名由 format 决定）
        format: 格式（csv / json / xlsx）

    Returns:
        实际输出文件路径

    Raises:
        ValueError: 不支持的格式
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    items = _build_bom_items(spec)

    if format == "csv":
        return _export_bom_csv(items, output_path)
    if format == "json":
        return _export_bom_json(items, output_path)
    if format == "xlsx":
        return _export_bom_xlsx(items, output_path)
    raise ValueError(f"不支持的 BOM 格式: {format}（支持 csv/json/xlsx）")


def _build_bom_items(spec: AssemblySpec) -> list[dict[str, Any]]:
    """从 AssemblySpec 构造 BOM 条目列表。"""
    items: list[dict[str, Any]] = []
    for idx, part in enumerate(spec.parts, start=1):
        quantity = part.quantity if part.quantity > 0 else 1
        mass = part.mass if part.mass is not None else 0.0
        total_mass = mass * quantity
        items.append({
            "item_number": idx,
            "part_id": part.part_id,
            "part_number": part.part_number or "",
            "name": part.name,
            "part_type": part.part_type,
            "quantity": quantity,
            "material": part.material or "",
            "mass": round(mass, 4),
            "total_mass": round(total_mass, 4),
            "remark": "",
        })
    return items


def _export_bom_csv(items: list[dict[str, Any]], path: Path) -> Path:
    """导出 CSV（UTF-8 with BOM，Excel 友好）。"""
    fieldnames = [
        "item_number", "part_number", "name", "part_type",
        "quantity", "material", "mass", "total_mass", "remark",
    ]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=fieldnames, extrasaction="ignore",
        )
        writer.writeheader()
        for item in items:
            writer.writerow(item)
    log.info("assembly.bom.csv_exported", path=str(path), items=len(items))
    return path


def _export_bom_json(items: list[dict[str, Any]], path: Path) -> Path:
    """导出 JSON。"""
    data = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "total_items": len(items),
        "total_quantity": sum(item["quantity"] for item in items),
        "total_mass": round(sum(item["total_mass"] for item in items), 4),
        "items": items,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log.info("assembly.bom.json_exported", path=str(path), items=len(items))
    return path


def _export_bom_xlsx(items: list[dict[str, Any]], path: Path) -> Path:
    """导出 Excel（需 openpyxl）。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as e:
        raise ValueError(
            "xlsx 格式需要 openpyxl，请安装: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # 标题行
    headers = [
        "件号", "图号", "名称", "类型",
        "数量", "材料", "单件质量(kg)", "总质量(kg)", "备注",
    ]
    ws.append(headers)

    # 标题样式
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for item in items:
        ws.append([
            item["item_number"], item["part_number"], item["name"],
            item["part_type"], item["quantity"], item["material"],
            item["mass"], item["total_mass"], item["remark"],
        ])
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    col_widths = [6, 20, 30, 10, 8, 15, 14, 14, 20]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    wb.save(path)
    log.info("assembly.bom.xlsx_exported", path=str(path), items=len(items))
    return path


# ===== 装配图（DXF）导出 =====


def export_assembly_drawing(
    spec: AssemblySpec,
    output_path: Path,
    paper_size: str = "A3",
) -> Path:
    """导出简化装配图（DXF）。

    简化策略：
    - 用零件 AABB 绘制矩形俯视图
    - 标注件号（引线 + 数字）
    - 绘制标题栏 + 明细栏
    - 不绘制详细几何（依赖 pythonOCC/CadQuery 才能生成精确视图）

    Args:
        spec: 装配规范
        output_path: 输出 DXF 文件路径
        paper_size: 图纸幅面（A0/A1/A2/A3/A4，默认 A3）

    Returns:
        实际输出 DXF 文件路径
    """
    try:
        import ezdxf
    except ImportError as e:
        raise ValueError("装配图导出需要 ezdxf，请安装: pip install ezdxf") from e

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    doc = ezdxf.new("R2010", setup=True)
    msp = doc.modelspace()

    # 图纸幅面尺寸（mm）
    paper_sizes = {
        "A0": (1189, 841), "A1": (841, 594), "A2": (594, 420),
        "A3": (420, 297), "A4": (297, 210),
    }
    paper_w, paper_h = paper_sizes.get(paper_size, paper_sizes["A3"])
    margin = 10  # 图框边距

    # 图框
    msp.add_lwpolyline([
        (margin, margin), (paper_w - margin, margin),
        (paper_w - margin, paper_h - margin), (margin, paper_h - margin),
        (margin, margin),
    ], close=True, dxfattribs={"layer": "图框", "lineweight": 50})

    # 计算零件 AABB（基于参数估计）
    from app.services.assembly.validator import _estimate_part_aabb
    from app.services.assembly.mate_library import (
        apply_mate_transforms, _list_to_mat,
    )
    transforms, _ = apply_mate_transforms(spec.parts, spec.mates)

    # 视图区域（左上角）
    view_x0 = margin + 10
    view_y0 = paper_h - margin - 10
    view_w = paper_w - 2 * margin - 20
    view_h = paper_h - 2 * margin - 100  # 留出底部标题栏空间

    # 计算所有零件的世界 AABB 包围盒
    import numpy as np
    all_min = [float("inf")] * 3
    all_max = [float("-inf")] * 3
    part_aabbs: list[tuple[str, tuple]] = []
    for part in spec.parts:
        local_aabb = _estimate_part_aabb(part)
        t = transforms.get(part.part_id, [
            1.0, 0.0, 0.0, 0.0, 0.0, 1.0, 0.0, 0.0,
            0.0, 0.0, 1.0, 0.0, 0.0, 0.0, 0.0, 1.0,
        ])
        t_mat = _list_to_mat(t)
        # 应用变换到 AABB 角点
        corners = []
        for x in (local_aabb[0][0], local_aabb[1][0]):
            for y in (local_aabb[0][1], local_aabb[1][1]):
                for z in (local_aabb[0][2], local_aabb[1][2]):
                    p = t_mat @ np.array([x, y, z, 1.0])
                    corners.append(p[:3])
        corners_arr = np.array(corners)
        w_min = corners_arr.min(axis=0)
        w_max = corners_arr.max(axis=0)
        part_aabbs.append((part.part_id, (w_min, w_max)))
        for i in range(3):
            all_min[i] = min(all_min[i], w_min[i])
            all_max[i] = max(all_max[i], w_max[i])

    bbox_w = max(all_max[0] - all_min[0], 1.0)
    bbox_h = max(all_max[1] - all_min[1], 1.0)
    scale = min(view_w / bbox_w, view_h / bbox_h) * 0.9

    def _to_paper(x: float, y: float) -> tuple[float, float]:
        """世界坐标 → 图纸坐标（保持比例，居中）。"""
        px = view_x0 + (x - all_min[0]) * scale + (view_w - bbox_w * scale) / 2
        py = view_y0 - (y - all_min[1]) * scale - (view_h - bbox_h * scale) / 2
        return px, py

    # 绘制每个零件的 AABB 矩形（俯视图：XY 平面投影）
    for idx, (pid, (w_min, w_max)) in enumerate(part_aabbs, start=1):
        x0, y0 = _to_paper(w_min[0], w_min[1])
        x1, y1 = _to_paper(w_max[0], w_max[1])
        msp.add_lwpolyline([
            (x0, y0), (x1, y0), (x1, y1), (x0, y1), (x0, y0),
        ], close=True, dxfattribs={"layer": "零件轮廓"})
        # 件号标注（中心位置）
        cx = (x0 + x1) / 2
        cy = (y0 + y1) / 2
        msp.add_text(
            str(idx), dxfattribs={
                "layer": "件号", "height": 5,
                "insert": (cx, cy),
            }
        )

    # 标题栏（右下角，GB/T 4458.2 风格简化版）
    title_w = 180
    title_h = 56
    title_x0 = paper_w - margin - title_w
    title_y0 = margin
    msp.add_lwpolyline([
        (title_x0, title_y0),
        (title_x0 + title_w, title_y0),
        (title_x0 + title_w, title_y0 + title_h),
        (title_x0, title_y0 + title_h),
        (title_x0, title_y0),
    ], close=True, dxfattribs={"layer": "标题栏"})

    # 标题栏内容
    msp.add_text(
        spec.name, dxfattribs={
            "layer": "标题栏", "height": 5,
            "insert": (title_x0 + 10, title_y0 + title_h - 15),
        }
    )
    msp.add_text(
        f"版本: {spec.version}", dxfattribs={
            "layer": "标题栏", "height": 3,
            "insert": (title_x0 + 10, title_y0 + title_h - 30),
        }
    )
    msp.add_text(
        f"单位: {spec.units}", dxfattribs={
            "layer": "标题栏", "height": 3,
            "insert": (title_x0 + 10, title_y0 + title_h - 40),
        }
    )
    msp.add_text(
        f"日期: {time.strftime('%Y-%m-%d')}", dxfattribs={
            "layer": "标题栏", "height": 3,
            "insert": (title_x0 + 100, title_y0 + title_h - 30),
        }
    )

    # 明细栏（标题栏上方，从下到上）
    bom_items = _build_bom_items(spec)
    bom_row_h = 8
    bom_rows = len(bom_items)
    bom_x0 = title_x0
    bom_y0 = title_y0 + title_h
    # 明细栏边框
    msp.add_lwpolyline([
        (bom_x0, bom_y0),
        (bom_x0 + title_w, bom_y0),
        (bom_x0 + title_w, bom_y0 + bom_row_h * bom_rows),
        (bom_x0, bom_y0 + bom_row_h * bom_rows),
        (bom_x0, bom_y0),
    ], close=True, dxfattribs={"layer": "明细栏"})
    # 明细栏行
    for i, item in enumerate(bom_items):
        row_y = bom_y0 + bom_row_h * (bom_rows - 1 - i)
        # 横线
        msp.add_line(
            (bom_x0, row_y), (bom_x0 + title_w, row_y),
            dxfattribs={"layer": "明细栏"},
        )
        # 件号 + 名称 + 数量
        msp.add_text(
            f"{item['item_number']}. {item['name']} ×{item['quantity']}",
            dxfattribs={
                "layer": "明细栏", "height": 3,
                "insert": (bom_x0 + 5, row_y + 2),
            }
        )

    doc.saveas(str(output_path))
    log.info(
        "assembly.drawing.exported",
        path=str(output_path),
        paper_size=paper_size,
        parts=len(spec.parts),
    )
    return output_path


def _np_array_ish(values):
    """辅助：将列表转为 numpy 数组（避免全局 import 冲突）。"""
    import numpy as _np
    return _np.array(values, dtype=_np.float64)


# ===== 模块自检 =====


def _self_test() -> dict[str, Any]:
    """离线自检：BOM 导出 + 装配图导出。"""
    from app.services.assembly.standard_parts import create_part
    from app.schemas.assembly import MateSpec

    checks: dict[str, bool] = {}
    errors: list[str] = []

    try:
        import ezdxf  # noqa: F401
        checks["ezdxf_available"] = True
    except ImportError:
        checks["ezdxf_available"] = False
        errors.append("ezdxf 不可用，装配图导出测试将跳过")

    # 构造测试装配体
    parts = [
        create_part("flange_plate", "flange-001", {
            "outer_diameter": 100.0, "inner_diameter": 50.0, "thickness": 10.0,
        }, quantity=2),
        create_part("bolt_iso4762", "bolt-001", {"m": 8.0, "length": 30.0}, quantity=6),
        create_part("bearing_6200", "bearing-001", {
            "outer_diameter": 28.0, "inner_diameter": 12.0, "width": 8.0,
        }, quantity=2),
    ]
    mates = [
        MateSpec(
            name="bolt_on_flange", type="coincident",
            part_a_id="flange-001", port_a_name="flange_face_b",
            part_b_id="bolt-001", port_b_name="head_bottom_face",
        ),
        MateSpec(
            name="bearing_on_flange", type="concentric",
            part_a_id="flange-001", port_a_name="center_axis",
            part_b_id="bearing-001", port_b_name="outer_ring_axis",
        ),
    ]
    spec = AssemblySpec(
        name="测试装配体", parts=parts, mates=mates,
        axioms=[], description="自检测试",
    )

    # BOM CSV
    import tempfile
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        try:
            csv_path = export_bom(spec, tmp_path / "bom.csv", "csv")
            checks["bom_csv_created"] = csv_path.is_file()
            checks["bom_csv_size_positive"] = csv_path.stat().st_size > 0
            # 读回验证
            with open(csv_path, "r", encoding="utf-8-sig") as f:
                content = f.read()
            checks["bom_csv_has_header"] = "件号" in content or "item_number" in content
            checks["bom_csv_has_items"] = content.count("\n") >= len(parts) + 1
        except Exception as e:  # noqa: BLE001
            errors.append(f"BOM CSV 失败: {e}")
            checks["bom_csv_created"] = False

        # BOM JSON
        try:
            json_path = export_bom(spec, tmp_path / "bom.json", "json")
            checks["bom_json_created"] = json_path.is_file()
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            checks["bom_json_has_items"] = "items" in data and len(data["items"]) == len(parts)
            checks["bom_json_total_quantity"] = (
                data["total_quantity"] == 2 + 6 + 2
            )
        except Exception as e:  # noqa: BLE001
            errors.append(f"BOM JSON 失败: {e}")
            checks["bom_json_created"] = False

        # 装配图 DXF
        if checks.get("ezdxf_available", False):
            try:
                dxf_path = export_assembly_drawing(
                    spec, tmp_path / "assembly.dxf", "A3",
                )
                checks["dxf_created"] = dxf_path.is_file()
                checks["dxf_size_positive"] = dxf_path.stat().st_size > 0
                # 读回验证
                import ezdxf
                doc = ezdxf.readfile(str(dxf_path))
                msp = doc.modelspace()
                checks["dxf_has_entities"] = len(msp) > 0
                layers = set(e.dxf.layer for e in msp)
                checks["dxf_has_layers"] = "图框" in layers and "标题栏" in layers
            except Exception as e:  # noqa: BLE001
                errors.append(f"DXF 装配图失败: {e}")
                checks["dxf_created"] = False

    # 不支持的格式报错
    try:
        export_bom(spec, Path("/tmp/x.txt"), "pdf")
        checks["unsupported_format_errors"] = False
    except ValueError:
        checks["unsupported_format_errors"] = True

    ok = all(checks.values())
    return {"ok": ok, "errors": errors, "checks": checks}


if __name__ == "__main__":  # pragma: no cover
    import sys
    result = _self_test()
    print(json.dumps(result, ensure_ascii=False, indent=2))
    sys.exit(0 if result["ok"] else 1)
